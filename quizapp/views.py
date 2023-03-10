from django.shortcuts import render,redirect,HttpResponse
from.models import *
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from.forms import *
from openpyxl import *


@login_required
def index(request):
        quiz_list=Quiz.objects.filter(private=False)
        context={
            'quiz_list': quiz_list
        }
        return render(request,'quiz/index.html',context)

def edit_index(request):
    if request.user.has_perm('quizapp.change_quiz'):
        quiz_list=Quiz.objects.all()
        context={
            'quiz_list': quiz_list
        }
        return render(request,'quiz/index_edit.html',context)
    else:
        return redirect('index')

@login_required
def leaderboards_index(request):
    quiz_list=Quiz.objects.all()
    context={
            'quiz_list': quiz_list
        }
    return render(request,'quiz/leaderboards_index.html',context)

def export_index(request):
    if request.user.has_perm('quizapp.add_scq'):
        quiz_list=Quiz.objects.all()
        context={
            'quiz_list': quiz_list
        }
        return render(request,'quiz/exportindex.html',context)

@login_required
def private_quiz_index(request):
    quiz_list=Quiz.objects.filter(private=True)
    context={
            'quiz_list': quiz_list
        }
    return render(request,'quiz/private_index.html',context)

@login_required
def private_quiz(request,quiz_name1):
    if request.method=='POST':
        password=request.POST.get('password')
        quiz1=Quiz.objects.get(quiz_name=quiz_name1)
        # privatequiz=Quiz.objects.filter(quiz=quiz1)
        if password==quiz1.password:
            return redirect(f'/quiz/{quiz1.quiz_name}/')
        else:
            return redirect('private_quiz_index')    
    else:
        return render(request,'quiz/privatequizpass.html')   

@login_required
def quiz(request,quiz_name1):
    quiz1=Quiz.objects.get(quiz_name=quiz_name1)
    question_scq=Scq.objects.filter(quiz=quiz1)
    question_int=intques.objects.filter(quiz=quiz1)
    question_boolean=booleanques.objects.filter(quiz=quiz1)
    question_mcq=Mcq.objects.filter(quiz=quiz1)
    
    if request.method=='POST':
        correct1=0
        wrong1=0
        marks_scored=0
        total=4*(len(question_boolean)+len(question_int)+len(question_scq)+len(question_mcq))
        for q in question_scq:
            
            if request.POST.get(q.question_text) != str(q.answer):
                wrong1+=1
                if q.marking=='nonnegmarking':
                    marks_scored-=0
                elif q.marking=='negmarking':
                    marks_scored-=1
            elif  request.POST.get(q.question_text) == str(q.answer):
                correct1+=1
                marks_scored+=4 
        for q in question_int:
            if int(request.POST.get(q.question_text)) != q.answer:
                wrong1+=1
                if q.marking=='nonnegmarking':
                    marks_scored-=0
                elif q.marking=='negmarking':
                    marks_scored-=1
            elif  int(request.POST.get(q.question_text))  == q.answer:
                correct1+=1
                marks_scored+=4 
        for q in question_boolean:
            if request.POST.get(q.question_text) != q.answer:
                wrong1=wrong1+1
                if q.marking=='nonnegmarking':
                    marks_scored-=0
                elif q.marking=='negmarking':
                    marks_scored-=1
            elif  request.POST.get(q.question_text) == q.answer:
                correct1+=1
                marks_scored+=4
        for q in question_mcq:
            ans_given=[
                request.POST.get(q.choice_1),
                request.POST.get(q.choice_2),
                request.POST.get(q.choice_3),
                request.POST.get(q.choice_4),
            ]
            awns=[
                q.choice_1_ans,
                q.choice_2_ans,
                q.choice_3_ans,
                q.choice_4_ans,   
            ]
            if ans_given[0]!=ans[0]:
                wrong1+=1
                if q.marking=='nonnegmarking':
                    marks_scored-=0
                elif q.marking=='negmarking':
                    marks_scored-=1   
            elif ans_given[1]!=ans[1]:
                wrong1+=1
                if q.marking=='nonnegmarking':
                    marks_scored-=0
                elif q.marking=='negmarking':
                    marks_scored-=1
            elif ans_given[2]!=ans[2]:
                wrong1+=1
                if q.marking=='nonnegmarking':
                    marks_scored-=0
                elif q.marking=='negmarking':
                    marks_scored-=1
            elif ans_given[3]!=ans[3]:
                wrong1+=1
                if q.marking=='nonnegmarking':
                    marks_scored-=0
                elif q.marking=='negmarking':
                    marks_scored-=1
            else:
                correct1+=1 
                marks_scored+=4    
        
        request.user.result_set.create(quiz=quiz1,correct=correct1,wrong=wrong1,marks_scored=marks_scored,maximum_marks=total,quiz_given=True)
        result=Result.objects.get(user=request.user,quiz=quiz1)
        context={
            'correct':result.correct,
            'wrong':result.wrong,
            'marks_scored':result.marks_scored,
            'maximum_marks':result.maximum_marks
        }
        
        return render(request,'quiz/result.html',context)
    else:
        if request.user.result_set.filter(quiz=quiz1,user=request.user):
            result=Result
            result=request.user.result_set.filter(quiz=quiz1,user=request.user).first()
            context={
            'correct':result.correct,
            'wrong':result.wrong,
            'marks_scored':result.marks_scored,
            'maximum_marks':result.maximum_marks
        }
            return  render(request,'quiz/result.html',context)
        else:
            context={
                'question_scq':question_scq,
                'question_int':question_int,
                'question_boolean':question_boolean,
                'question_mcq':question_mcq,                
                'quizz':quiz1
            }
            return render(request,'quiz/question.html',context)

def addquiz(request):
    if request.user.has_perm('quizapp.add_quiz'):
        form=addquizform(request.POST)
        if(request.method=='POST'):
            if(form.is_valid()):
                form.save()
                return redirect('addquestion')
        else:
            context={
                'form':form
            }
            return render(request,'quiz/addquiz.html',context)
    else:
        return redirect('index')

def editquestion(request,quiz_name1,question_type,id):
    if request.user.has_perm('quizapp.change_scq'):
        if request.method=='POST':
            if request.POST.get('question_type')=='scq':
                form=ScqForm(request.POST)
                form.save()
            elif request.POST.get('question_type')=='mcq':
                form=McqForm(request.POST)
                form.save()
            elif request.POST.get('question_type')=='int':
                form=IntForm(request.POST)
                form.save()
            elif request.POST.get('question_type')=='bool':
                form=BooleanForm(request.POST)
                form.save()
            return redirect('index')
        else:
            
            context={
                'question_type':question_type
            }
            if question_type=='scq':
                question=Scq.objects.get(id=id)
                form=ScqForm(instance=question) 
                context['form']=form 
                question.delete() 
                return render(request,'quiz/addquestion.html',context)
            
            elif question_type=='mcq':
                question=Mcq.objects.get(id=id)
                form=McqForm(instance=question)  
                context['form']=form 
                question.delete() 
                return render(request,'quiz/addquestion.html',context)
            
            elif question_type=='int':
                question=intques.objects.get(id=id)
                form=IntForm(instance=question)  
                context['form']=form 
                question.delete() 
                return render(request,'quiz/addquestion.html',context)
            elif question_type=='bool':
                question=booleanques.objects.get(id=id)
                form=BooleanForm(instance=question)  
                context['form']=form  
                question.delete()
                return render(request,'quiz/addquestion.html',context)
            else:
                return redirect('index')

def edit_quiz(request,quiz_name1):
    if request.user.has_perm('quizapp.change_question'):
        quiz1=Quiz.objects.get(quiz_name=quiz_name1)
        question_scq=Scq.objects.filter(quiz=quiz1)
        question_mcq=Mcq.objects.filter(quiz=quiz1)
        question_int=intques.objects.filter(quiz=quiz1)
        question_bool=booleanques.objects.filter(quiz=quiz1)
        
        context={
                    'question_scq':question_scq,
                    'question_mcq':question_mcq,
                    'question_int':question_int,
                    'question_bool':question_bool,
                    'quizz':quiz1
                }
        return render(request,'quiz/quiz_edit.html',context)
    else:
        return redirect('index')

@login_required
def leaderboards(request,quiz_name1):
    quiz1=Quiz.objects.get(quiz_name=quiz_name1)
    result_list=Result.objects.filter(quiz=quiz1).order_by('-marks_scored')
    context={
        "result_list":result_list
    }
    return render(request,'quiz/leaderboards.html',context)

def question_type(request):
    if request.user.has_perm('quizapp.change_question'):
        if request.method=='POST':
            if request.POST.get('question_type')=='scq':
                form=ScqForm()
            elif request.POST.get('question_type')=='mcq':
                form=McqForm()
            elif request.POST.get('question_type')=='int':
                form=IntForm()
            elif request.POST.get('question_type')=='bool':  
                form=BooleanForm()
            context={
                'form':form,
                'question_type':request.POST.get('question_type')
            }   
            return render(request,'quiz/addquestion.html',context)    
        return render(request,'quiz/questiontype.html')   
def addQuestion(request):    
    if request.user.has_perm('quizapp.add_scq'):
        if request.method=='POST':
            if request.POST.get('question_type')=='scq':
                form=ScqForm(request.POST)
                form.save()
            elif request.POST.get('question_type')=='mcq':
                form=McqForm(request.POST)
                form.save()
            elif request.POST.get('question_type')=='int':
                form=IntForm(request.POST)
                form.save()
            elif request.POST.get('question_type')=='bool':
                form=BooleanForm(request.POST)
                form.save()
            
            
        return redirect('index')
        
            
    else: 
        return redirect('index')
def export_result(request,quiz_name1):
    if request.user.has_perm('quizapp.add_scq'):
        wb =Workbook()
        ws1=wb.active
        quiz1=Quiz.objects.get(quiz_name=quiz_name1)
        
        resultlist=Result.objects.filter(quiz=quiz1)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="result.xlsx"'
        num=1
        columns = [
            'User',
            'Result',
        ]
        
        for col_num, column_title in enumerate(columns, 1):
            cell = ws1.cell(row=num, column=col_num)
            cell.value = column_title
        for result in resultlist:
            num+=1
            user=str(result.user)
            marks=int(result.marks_scored)
            row=[
                user,
                marks
            ]
            for col_num, cell_value in enumerate(row, 1):
                cell = ws1.cell(row=num, column=col_num)
                cell.value = cell_value
        wb.save(response)
        return response


